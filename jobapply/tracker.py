"""Excel application tracker. One row per job; updated as it moves through the
pipeline so you always have a shareable, sortable record of every application."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import Job

SHEET = "Applications"

# (header, Job attribute, column width)
COLUMNS = [
    ("Date", "captured_at", 20),
    ("Company", "company", 24),
    ("Role", "title", 30),
    ("Location", "location", 22),
    ("Status", "status", 16),
    ("Fit", "fit_score", 8),
    ("Why it fits", "fit_reasons", 50),
    ("Job URL", "url", 45),
    ("Resume PDF", "pdf_path", 40),
    ("Job ID", "id", 18),
]

_HEADER_FILL = PatternFill("solid", fgColor="20242C")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


class Tracker:
    def __init__(self, xlsx_path: Path):
        self.path = xlsx_path
        if not self.path.exists():
            self._create()

    def _create(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET
        for col_idx, (header, _, width) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(vertical="center")
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.freeze_panes = "A2"
        self.path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(self.path)

    def _find_row(self, ws, job_id: str) -> int | None:
        id_col = len(COLUMNS)  # "Job ID" is the last column
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=id_col).value == job_id:
                return r
        return None

    def upsert(self, job: Job) -> None:
        """Insert the job as a new row, or update its existing row in place."""
        wb = load_workbook(self.path)
        ws = wb[SHEET]
        row = self._find_row(ws, job.id)
        if row is None:
            row = ws.max_row + 1
        data = job.to_row()
        for col_idx, (_, attr, _) in enumerate(COLUMNS, start=1):
            value = data.get(attr, "")
            ws.cell(row=row, column=col_idx,
                    value="" if value is None else value).alignment = Alignment(
                        wrap_text=(attr in {"fit_reasons"}), vertical="top")
        wb.save(self.path)
